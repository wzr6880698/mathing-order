import pandas as pd
import streamlit as st
import re
import io
from datetime import datetime
import openpyxl
import numpy as np


def col_num_to_letter(n):
    """将 0-based 列索引转换为 Excel 列字母"""
    letters = []
    while n >= 0:
        letters.append(chr(65 + (n % 26)))
        n = n // 26 - 1
    return ''.join(reversed(letters))


def like_order_string(s):
    """判断一个字符串是否像订单号"""
    if pd.isna(s) or not isinstance(s, str):
        return False
    if len(s) < 3 or len(s) > 50:
        return False
    # 包含中文字符的通常不是订单号（内销订单号可能包含特定前缀，保留此规则）
    if re.search('[\u4e00-\u9fff]', s):
        return False
    # 允许字母、数字、短横线、下划线、斜杠、点
    if re.match(r'^[A-Za-z0-9\-_/\.]+$', s):
        return True
    return False


def is_valid_header_cell(cell_value):
    """判断单元格值是否为有效表头（非空、非纯数字、非订单号格式）"""
    if pd.isna(cell_value):
        return False
    cell_str = str(cell_value).strip()
    # 空字符串不是有效表头
    if cell_str == '':
        return False
    # 纯数字（无单位）大概率不是表头
    if re.match(r'^\d+(\.\d+)?$', cell_str) and not like_order_string(cell_str):
        return False
    # 订单号格式不是表头
    if like_order_string(cell_str):
        return False
    # 长度超过50的大概率不是表头
    if len(cell_str) > 50:
        return False
    return True


def auto_detect_header_row(file_obj):
    """
    智能识别Excel中的真实表头行（核心功能）
    返回：(header_row_index: int, header_names: list)
    - header_row_index: 表头行的索引（0-based），None表示未找到
    - header_names: 表头名称列表
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    max_check_rows = 20  # 最多检查前20行，覆盖大部分标题场景
    header_candidates = []

    # 遍历前20行，评估每行作为表头的可能性
    for row_idx in range(min(max_check_rows, ws.max_row)):
        row_cells = ws[row_idx + 1]  # openpyxl是1-based
        row_values = [cell.value for cell in row_cells]

        # 计算该行的有效表头单元格数量
        valid_header_count = sum(1 for val in row_values if is_valid_header_cell(val))
        total_non_empty = sum(1 for val in row_values if val is not None and str(val).strip() != '')

        # 有效表头占比 = 有效表头数 / 非空单元格数（避免全空行）
        if total_non_empty == 0:
            header_score = 0
        else:
            header_score = valid_header_count / total_non_empty

        # 记录候选行（行索引、得分、值列表）
        header_candidates.append({
            'row_idx': row_idx,
            'score': header_score,
            'values': row_values
        })

    # 找到得分最高的行作为表头行
    if not header_candidates:
        return 0, ['列1', '列2', '列3']  # 兜底

    best_header = max(header_candidates, key=lambda x: x['score'])
    header_row_idx = best_header['row_idx']
    header_values = best_header['values']

    # 清理表头名称
    clean_headers = []
    for val in header_values:
        if val is None:
            # 空表头用列位置命名（如A、B、C）
            col_idx = len(clean_headers)
            col_letter = col_num_to_letter(col_idx)
            clean_headers.append(f'{col_letter}列')
        else:
            # 清理特殊字符和重复
            clean_val = str(val).strip().replace('\n', ' ').replace('\t', ' ')
            clean_val = re.sub(r'[^\w\s\u4e00-\u9fff]', '', clean_val)
            if clean_val == '':
                col_idx = len(clean_headers)
                col_letter = col_num_to_letter(col_idx)
                clean_headers.append(f'{col_letter}列')
            else:
                # 去重处理
                if clean_val in clean_headers:
                    count = clean_headers.count(clean_val) + 1
                    clean_val = f'{clean_val}_{count}'
                clean_headers.append(clean_val)

    return header_row_idx, clean_headers


def read_excel_with_auto_header(file_obj):
    """
    智能读取Excel文件（自动识别表头行）
    返回：处理后的DataFrame
    """
    # 第一步：识别表头行
    header_row_idx, header_names = auto_detect_header_row(file_obj)

    # 重置文件指针（避免读取失败）
    file_obj.seek(0)

    # 第二步：读取数据（跳过表头行之前的行）
    df = pd.read_excel(
        file_obj,
        header=None,
        skiprows=header_row_idx + 1,  # skiprows是跳过的行数，表头行本身不跳过
        engine='openpyxl'
    )

    # 第三步：设置表头名称（截断/补全以匹配列数）
    df_cols = len(df.columns)
    if len(header_names) > df_cols:
        header_names = header_names[:df_cols]
    elif len(header_names) < df_cols:
        # 补全列名
        for i in range(len(header_names), df_cols):
            col_letter = col_num_to_letter(i)
            header_names.append(f'{col_letter}列')

    df.columns = header_names

    # 第四步：清理数据（删除全空行/列）
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')

    # 重置索引
    df = df.reset_index(drop=True)

    return df, header_row_idx


def fill_merged_cells(df, order_col):
    """填充合并单元格导致的空值（针对订单号列）"""
    if order_col in df.columns:
        # 先将NaN转为空字符串，再向前填充
        df[order_col] = df[order_col].fillna('').astype(str)
        df[order_col] = df[order_col].replace('', method='ffill')
    return df


def detect_domestic_order_column(df):
    """智能识别内销订单号列"""
    # 内销订单号列关键词（按优先级排序）
    domestic_keywords = [
        '订单号', '内销订单号', '订单编号', '内销编号', '单号', '内销单号',
        '编号', '序号', 'ID', 'id', 'No', 'NO', 'no', 'A列', 'B列', 'C列'
    ]

    # 遍历关键词找匹配列
    for keyword in domestic_keywords:
        for col in df.columns:
            col_str = str(col).lower()
            if keyword.lower() in col_str:
                return col

    # 如果没有匹配，返回非空值最多的列（订单号列通常非空值多）
    non_null_counts = df.notna().sum()
    if not non_null_counts.empty:
        return non_null_counts.idxmax()

    # 最后返回第一列（内销订单号通常在首列）
    return df.columns[0] if len(df.columns) > 0 else None


def detect_export_order_column(df):
    """智能识别外销订单号列"""
    export_keywords = [
        '订单号', '订单编号', 'Order Number', 'order number',
        'orderno', 'Order No', 'order no', '订单ID'
    ]

    for keyword in export_keywords:
        for col in df.columns:
            if keyword.lower() in str(col).lower():
                return col

    non_null_counts = df.notna().sum()
    if not non_null_counts.empty:
        return non_null_counts.idxmax()

    return df.columns[0] if len(df.columns) > 0 else None


def detect_file_type_and_order_col(df, file_name):
    """综合识别：文件类型 + 推荐订单号列"""
    # 先通过文件名判断
    if any(keyword in file_name.lower() for keyword in ['内销', 'domestic']):
        return 'domestic', detect_domestic_order_column(df)
    elif any(keyword in file_name.lower() for keyword in ['外销', 'export']):
        return 'export', detect_export_order_column(df)

    # 文件名无标识，通过内容判断
    col_names = ' '.join([str(col).lower() for col in df.columns])
    domestic_score = sum(1 for kw in ['内销', '国内'] if kw in col_names)
    export_score = sum(1 for kw in ['外销', '出口', 'export'] if kw in col_names)

    if domestic_score > export_score:
        return 'domestic', detect_domestic_order_column(df)
    else:
        return 'export', detect_export_order_column(df)


def main():
    st.set_page_config(
        page_title="智能订单匹配工具",
        page_icon="🔗",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    # 侧边栏说明
    with st.sidebar:
        st.header("📖 智能匹配工具 - 使用说明")
        st.markdown("""
        ### 🚀 核心优势
        ✅ **全自动表头识别**：无需固定跳过行数，适配任意标题格式
        ✅ **智能列名展示**：无Unnamed列名，显示真实友好的列名
        ✅ **内销/外销兼容**：自动识别文件类型，无需手动切换
        ✅ **合并单元格处理**：自动填充内销明细表空订单号
        ✅ **订单号格式保护**：文本格式，避免科学计数法

        ### 📝 操作步骤
        1. 上传汇总表（内销/外销Excel）
        2. 上传明细表（内销/外销Excel）
        3. 确认系统推荐的订单号列（可手动调整）
        4. 点击匹配，查看智能分析结果
        5. 下载完整匹配数据

        ### 💡 适配场景
        - 标题占任意行数的Excel文件
        - 内销/外销不同格式的文件
        - 包含合并单元格的明细表
        - 列名不规范的Excel文件
        """)
        st.markdown("---")
        st.info("版本: v4.0.0（全自动智能表头识别）")

    # 主界面
    st.title("🔗 智能订单匹配工具（全自动表头识别）")
    st.markdown("### 适配任意格式 | 无需手动调整行数 | 内销/外销全覆盖")
    st.markdown("---")

    # 文件上传区域
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📊 汇总表")
        summary_file = st.file_uploader(
            "上传汇总Excel文件（支持任意标题行数）",
            type=['xlsx', 'xls'],
            key="summary_file",
            help="自动识别表头行，无需手动调整"
        )
        if summary_file:
            st.success(f"✅ 已上传: {summary_file.name}")

    with col2:
        st.subheader("📋 明细表")
        detail_file = st.file_uploader(
            "上传明细Excel文件（支持合并单元格）",
            type=['xlsx', 'xls'],
            key="detail_file",
            help="内销文件自动填充合并单元格空值"
        )
        if detail_file:
            st.success(f"✅ 已上传: {detail_file.name}")

    # 数据读取和预处理（核心：自动识别表头）
    df_summary = None
    df_detail = None
    summary_type = None
    detail_type = None
    summary_recommend_col = None
    detail_recommend_col = None
    summary_header_row = 0
    detail_header_row = 0

    # 读取汇总表
    if summary_file:
        try:
            # 关键：智能读取（自动识别表头）
            df_summary, summary_header_row = read_excel_with_auto_header(summary_file)
            st.info(f"""
            📊 汇总表分析结果：
            - 总行数：{len(df_summary)} 行
            - 总列数：{len(df_summary.columns)} 列
            - 自动识别表头行：第 {summary_header_row + 1} 行
            """)

            # 识别文件类型和推荐列
            summary_type, summary_recommend_col = detect_file_type_and_order_col(
                df_summary, summary_file.name
            )
            type_text = "内销" if summary_type == "domestic" else "外销"
            st.success(f"📌 文件类型：{type_text}格式")
            if summary_recommend_col:
                st.info(f"💡 推荐订单号列：**{summary_recommend_col}**")

        except Exception as e:
            st.error(f"❌ 读取汇总表失败：{str(e)}")
            st.warning("建议检查文件格式，确保是标准Excel文件（.xlsx/.xls）")

    # 读取明细表
    if detail_file:
        try:
            # 关键：智能读取（自动识别表头）
            df_detail, detail_header_row = read_excel_with_auto_header(detail_file)
            st.info(f"""
            📋 明细表分析结果：
            - 总行数：{len(df_detail)} 行
            - 总列数：{len(df_detail.columns)} 列
            - 自动识别表头行：第 {detail_header_row + 1} 行
            """)

            # 识别文件类型和推荐列
            detail_type, detail_recommend_col = detect_file_type_and_order_col(
                df_detail, detail_file.name
            )
            type_text = "内销" if detail_type == "domestic" else "外销"
            st.success(f"📌 文件类型：{type_text}格式")

            # 内销明细表自动填充合并单元格空值
            if detail_type == "domestic" and detail_recommend_col:
                df_detail = fill_merged_cells(df_detail, detail_recommend_col)
                st.success(f"✅ 已填充【{detail_recommend_col}】列合并单元格空值")

            if detail_recommend_col:
                st.info(f"💡 推荐订单号列：**{detail_recommend_col}**")

        except Exception as e:
            st.error(f"❌ 读取明细表失败：{str(e)}")
            st.warning("建议检查文件格式，确保是标准Excel文件（.xlsx/.xls）")

    st.markdown("---")

    # 列选择区域（友好列名展示）
    if df_summary is not None and df_detail is not None:
        st.subheader("🏷️ 订单号列选择（系统智能推荐）")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**汇总表（{summary_type}格式）：**")
            # 列选择框：显示真实友好的列名
            summary_col = st.selectbox(
                "选择汇总表订单号列",
                options=df_summary.columns.tolist(),
                index=df_summary.columns.tolist().index(summary_recommend_col)
                if summary_recommend_col in df_summary.columns else 0,
                key="summary_col",
                help="系统已根据内容推荐最优列"
            )

        with col2:
            st.markdown(f"**明细表（{detail_type}格式）：**")
            # 列选择框：显示真实友好的列名
            detail_col = st.selectbox(
                "选择明细表订单号列",
                options=df_detail.columns.tolist(),
                index=df_detail.columns.tolist().index(detail_recommend_col)
                if detail_recommend_col in df_detail.columns else 0,
                key="detail_col",
                help="内销文件已自动填充合并单元格空值"
            )

        # 手动填充按钮（内销专用）
        if detail_type == "domestic":
            st.markdown("---")
            if st.button("🔄 重新填充明细表订单号列空值", use_container_width=True):
                df_detail = fill_merged_cells(df_detail, detail_col)
                st.success(f"✅ 已重新填充【{detail_col}】列的空值（合并单元格）")

        # 显示列选择确认信息
        st.markdown("---")
        st.info(f"""
        📌 匹配配置确认：
        - 汇总表表头行：第 {summary_header_row + 1} 行 | 订单号列：**{summary_col}**
        - 明细表表头行：第 {detail_header_row + 1} 行 | 订单号列：**{detail_col}**
        - 文件类型：汇总表({summary_type}) | 明细表({detail_type})
        """)

        # 匹配按钮
        if st.button("🔗 开始智能匹配", type="primary", use_container_width=True):
            with st.spinner("正在执行智能匹配，请稍候..."):
                try:
                    # 数据预处理
                    # 1. 确保订单号列为字符串格式
                    df_summary[summary_col] = df_summary[summary_col].astype(str).str.strip()
                    df_detail[detail_col] = df_detail[detail_col].astype(str).str.strip()

                    # 2. 过滤无效订单号（空字符串/纯空格）
                    valid_summary_orders = df_summary[
                        df_summary[summary_col] != ''
                        ][summary_col].unique()
                    order_set = set(valid_summary_orders)

                    # 3. 执行匹配
                    matched_df = df_detail[df_detail[detail_col].isin(order_set)]

                    # 结果展示
                    if len(matched_df) == 0:
                        st.warning("⚠️ 未找到匹配的订单记录！")
                        with st.expander("📈 匹配分析（帮助排查问题）"):
                            st.markdown(f"- 汇总表有效订单数：{len(valid_summary_orders)}")
                            st.markdown(f"- 明细表有效记录数：{len(df_detail[df_detail[detail_col] != ''])}")
                            st.markdown(
                                f"- 汇总表订单号样本：{list(valid_summary_orders[:5]) if len(valid_summary_orders) > 0 else '无'}")
                            st.markdown(
                                f"- 明细表订单号样本：{list(df_detail[df_detail[detail_col] != ''][detail_col].unique()[:5]) if len(df_detail) > 0 else '无'}")
                            st.markdown(
                                f"- 汇总表表头行：第 {summary_header_row + 1} 行 | 明细表表头行：第 {detail_header_row + 1} 行")
                    else:
                        st.success(f"✅ 匹配完成！共找到 **{len(matched_df)}** 条匹配记录")

                        # 统计信息
                        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
                        with stat_col1:
                            st.metric("汇总表订单数", len(valid_summary_orders))
                        with stat_col2:
                            st.metric("明细表总记录数", len(df_detail))
                        with stat_col3:
                            st.metric("匹配记录数", len(matched_df))
                        with stat_col4:
                            valid_detail = len(df_detail[df_detail[detail_col] != ''])
                            match_rate = (len(matched_df) / valid_detail) * 100 if valid_detail > 0 else 0
                            st.metric("匹配率", f"{match_rate:.1f}%")

                        # 结果预览
                        st.subheader("📊 匹配结果预览")
                        st.dataframe(matched_df.head(30), use_container_width=True)
                        if len(matched_df) > 30:
                            st.caption(f"仅展示前30条，共{len(matched_df)}条匹配记录")

                        # 生成下载文件
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            matched_df.to_excel(writer, index=False, sheet_name='匹配结果')
                            # 设置订单号列为文本格式
                            workbook = writer.book
                            worksheet = writer.sheets['匹配结果']
                            col_idx = matched_df.columns.get_loc(detail_col)
                            col_letter = col_num_to_letter(col_idx)
                            text_format = workbook.add_format({'num_format': '@'})
                            worksheet.set_column(f'{col_letter}:{col_letter}', None, text_format)

                        output.seek(0)

                        # 下载按钮
                        st.markdown("---")
                        st.subheader("📥 下载匹配结果")
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        st.download_button(
                            label="📦 下载完整匹配结果 (Excel)",
                            data=output,
                            file_name=f"智能匹配结果_{timestamp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        st.info(f"""
                        💡 匹配结果说明：
                        - 订单号列已设置为文本格式，避免科学计数法
                        - 内销文件合并单元格空值已填充
                        - 汇总表表头行：第 {summary_header_row + 1} 行
                        - 明细表表头行：第 {detail_header_row + 1} 行
                        """)

                except Exception as e:
                    st.error(f"❌ 匹配过程出错：{str(e)}")
                    with st.expander("查看详细错误信息"):
                        import traceback
                        st.code(traceback.format_exc())


if __name__ == "__main__":
    main()